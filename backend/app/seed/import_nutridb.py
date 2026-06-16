"""Import exportního balíčku z NutriDatabaze.cz do tabulky `ingredient`.

Export si stáhneš po (bezplatné) registraci na nutridatabaze.cz → Data ke
stažení. Formát se může lišit (CSV/XLSX, různé názvy sloupců), proto je
mapování sloupců volné přes argumenty. Použití:

    python -m app.seed.import_nutridb /cesta/export.csv \
        --col-name "Název potraviny v češtině" \
        --col-name-en "Název potraviny v angličtině" \
        --col-code "Kód potraviny" \
        --col-kcal "Energie [kcal]"

Bez --col-* se skript pokusí sloupce odhadnout podle běžných názvů.
"""
from __future__ import annotations

import argparse
import sys

from sqlalchemy import select

from ..db import SessionLocal, engine
from ..db import Base
from ..models import Ingredient

GUESS = {
    "name": ["nazev potraviny v cestine", "nazev cesky", "nazev", "potravina"],
    "name_en": ["nazev potraviny v anglictine", "nazev anglicky", "name"],
    "code": ["kod potraviny", "kod", "code"],
    "kcal": ["energie [kcal]", "energie kcal", "kcal", "energie"],
    "protein": ["bilkoviny", "protein"],
    "carbs": ["sacharidy", "carbs"],
    "fat": ["tuky", "fat"],
    "fiber": ["vlaknina", "fiber"],
}


def _norm(s: str) -> str:
    import unicodedata

    s = "".join(
        c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)
    )
    return s.lower().strip()


def _load_rows(path: str) -> list[dict]:
    if path.lower().endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h) if h is not None else "" for h in rows[0]]
        return [dict(zip(header, r)) for r in rows[1:]]
    # CSV – zkus pár oddělovačů
    import csv

    with open(path, encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        delim = ";" if sample.count(";") >= sample.count(",") else ","
        return list(csv.DictReader(f, delimiter=delim))


def _resolve(header: list[str], explicit: dict[str, str | None]) -> dict[str, str]:
    norm_map = {_norm(h): h for h in header}
    chosen: dict[str, str] = {}
    for field, candidates in GUESS.items():
        if explicit.get(field):
            chosen[field] = explicit[field]  # type: ignore[assignment]
            continue
        for cand in candidates:
            if cand in norm_map:
                chosen[field] = norm_map[cand]
                break
    return chosen


def _num(val) -> float | None:
    if val in (None, ""):
        return None
    try:
        return float(str(val).replace(",", ".").split()[0])
    except (ValueError, IndexError):
        return None


def main() -> int:
    ap = argparse.ArgumentParser(description="Import NutriDatabaze.cz exportu")
    ap.add_argument("path")
    for f in GUESS:
        ap.add_argument(f"--col-{f.replace('_', '-')}", dest=f, default=None)
    args = ap.parse_args()

    rows = _load_rows(args.path)
    if not rows:
        print("Prázdný soubor.", file=sys.stderr)
        return 1

    explicit = {f: getattr(args, f) for f in GUESS}
    cols = _resolve(list(rows[0].keys()), explicit)
    if "name" not in cols:
        print(f"Nenašel jsem sloupec s názvem. Sloupce: {list(rows[0].keys())}")
        return 1
    print(f"Mapování sloupců: {cols}")

    Base.metadata.create_all(engine)
    db = SessionLocal()
    inserted = updated = 0
    try:
        for row in rows:
            name = (row.get(cols["name"]) or "").strip()
            if not name:
                continue
            code = (row.get(cols.get("code", "")) or "") or None
            existing = None
            if code:
                existing = db.scalar(select(Ingredient).where(Ingredient.code == code))
            if existing is None:
                existing = db.scalar(
                    select(Ingredient).where(Ingredient.name_cs == name)
                )
            ing = existing or Ingredient(name_cs=name)
            ing.name_cs = name
            if "name_en" in cols:
                ing.name_en = (row.get(cols["name_en"]) or None)
            ing.code = code
            ing.kcal_100g = _num(row.get(cols.get("kcal", "")))
            ing.protein_100g = _num(row.get(cols.get("protein", "")))
            ing.carbs_100g = _num(row.get(cols.get("carbs", "")))
            ing.fat_100g = _num(row.get(cols.get("fat", "")))
            ing.fiber_100g = _num(row.get(cols.get("fiber", "")))
            ing.source = "nutridatabaze"
            if existing is None:
                db.add(ing)
                inserted += 1
            else:
                updated += 1
        db.commit()
    finally:
        db.close()
    print(f"Hotovo: {inserted} nových, {updated} aktualizovaných surovin.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
